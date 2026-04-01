import os
# from select import poll
import uuid
import base64
import io
from flask import Flask, render_template, request, \
                  redirect, session, url_for, flash, jsonify, send_file
import pytz
from config import Config
from functools import wraps
from models import get_db_connection, init_db
from datetime import datetime, timedelta, timezone
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()

# ── Create Flask app ──────────────────────────────────────
app = Flask(__name__)
app.config.from_object(Config)
app.permanent_session_lifetime = timedelta(minutes=30)
IST = timezone(timedelta(hours=5, minutes=30))


def to_ist(dt):
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST)


@app.template_filter("format_ist")
def format_ist(dt, fmt="%d %b %Y, %I:%M %p"):
    local_dt = to_ist(dt)
    return local_dt.strftime(fmt) if local_dt else ""


@app.template_filter("datetime_local_ist")
def datetime_local_ist(dt):
    local_dt = to_ist(dt)
    return local_dt.strftime("%Y-%m-%dT%H:%M") if local_dt else ""

# ── Uploads folder ────────────────────────────────────────
UPLOADS_FOLDER = os.path.join(
    os.path.dirname(__file__),
    'static', 'uploads'
)
os.makedirs(UPLOADS_FOLDER, exist_ok=True)

# ── Initialize database ───────────────────────────────────
with app.app_context():
    init_db()

# ── Register Blueprints ───────────────────────────────────
from routes.poll_routes import poll_bp
from routes.vote_routes import vote_bp
from routes.admin_routes import admin_bp

app.register_blueprint(poll_bp)
app.register_blueprint(vote_bp)
app.register_blueprint(admin_bp)

# ── Session permanent on every request ───────────────────
@app.before_request
def make_session_permanent():
    session.permanent = True

# ── Admin decorator ───────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return render_template('404.html'), 404
        return f(*args, **kwargs)
    return decorated

# ── Auth Routes ───────────────────────────────────────────
@app.route('/login')
def login():
    return render_template('login.html',
                           hide_navbar=True,
                           hide_footer=True)

@app.route('/faqs')
def faqs():
    return render_template('faqs.html')

# ── Dashboard ─────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    conn    = get_db_connection()
    cursor  = conn.cursor()

    # Total polls
    cursor.execute("""
        SELECT COUNT(*) as count FROM polls
        WHERE user_id = %s AND status = 1
    """, (user_id,))
    total_polls = cursor.fetchone()['count']

    # Active polls
    cursor.execute("""
        SELECT COUNT(*) as count FROM polls
        WHERE user_id = %s
        AND status = 1
        AND end_time > NOW()
    """, (user_id,))
    active_polls = cursor.fetchone()['count']

    # Total votes
    cursor.execute("""
        SELECT COUNT(*) as count FROM votes
        WHERE poll_id IN (
            SELECT id FROM polls WHERE user_id = %s
        )
    """, (user_id,))
    total_votes = cursor.fetchone()['count']

    # Recent polls
    cursor.execute("""
        SELECT
            p.id,
            p.question,
            p.end_time,
            p.share_token,
            p.user_id,
            COUNT(v.id) as vote_count,
            u.first_name,
            u.last_name,
            CASE
                WHEN p.end_time <= NOW()
                    THEN 'Expired'
                WHEN p.start_time > NOW()
                    THEN 'Not Started'
                ELSE 'Active'
            END as status
        FROM polls p
        LEFT JOIN votes v ON v.poll_id = p.id
        LEFT JOIN users u ON u.id = p.user_id
        WHERE p.status = 1
        AND p.user_id = %s
        GROUP BY p.id, p.question, p.end_time,
                 p.share_token, p.user_id,
                 p.start_time, u.first_name, u.last_name
        ORDER BY p.created_at DESC
        LIMIT 5
    """, (user_id,))
    recent_polls = cursor.fetchall()

    cursor.close()
    conn.close()

    active_pct  = round((active_polls / total_polls * 100)) \
                  if total_polls > 0 else 0
    ended_polls = total_polls - active_polls
    ended_pct   = round((ended_polls / total_polls * 100)) \
                  if total_polls > 0 else 0
    max_votes   = 100
    votes_pct   = min(
        round((total_votes / max_votes * 100)), 100
    )

    return render_template('dashboard.html',
                           active_page  = 'dashboard',
                           total_polls  = total_polls,
                           active_polls = active_polls,
                           total_votes  = total_votes,
                           active_pct   = active_pct,
                           ended_pct    = ended_pct,
                           votes_pct    = votes_pct,
                           recent_polls = recent_polls)


# ── My Polls ──────────────────────────────────────────────
@app.route('/dashboard/polls')
def my_polls():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id  = session['user_id']
    page     = request.args.get('page', 1, type=int)
    per_page = 8
    offset   = (page - 1) * per_page

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) FROM polls
        WHERE user_id = %s AND status = 1
    """, (user_id,))
    total_count = cursor.fetchone()['count']

    cursor.execute("""
        SELECT
            p.id,
            p.question,
            p.start_time,
            p.end_time,
            p.share_token,
            COUNT(v.id) as vote_count,
            CASE
                WHEN p.end_time <= NOW()
                    THEN 'Expired'
                WHEN p.start_time > NOW()
                    THEN 'Not Started'
                ELSE 'Active'
            END as status
        FROM polls p
        LEFT JOIN votes v ON v.poll_id = p.id
        WHERE p.user_id = %s
        AND p.status = 1
        GROUP BY p.id, p.question, p.start_time,
                 p.end_time, p.share_token
        ORDER BY p.created_at DESC
        LIMIT %s OFFSET %s
    """, (user_id, per_page, offset))
    polls = cursor.fetchall()

    cursor.close()
    conn.close()

    total_pages = (total_count + per_page - 1) // per_page

    return render_template('my_polls.html',
                           active_page  = 'polls',
                           polls        = polls,
                           current_page = page,
                           total_pages  = total_pages)


# ── Poll Detail ───────────────────────────────────────────
@app.route('/dashboard/poll/<string:token>')
def poll_detail(token):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM polls
        WHERE share_token = %s AND user_id = %s
    """, (token, session['user_id']))
    poll = cursor.fetchone()

    if not poll:
        cursor.close()
        conn.close()
        return render_template('404.html'), 404

    poll_id = poll["id"]

    cursor.execute("""
        SELECT
            o.id,
            o.option,
            COUNT(v.id) as vote_count
        FROM options o
        LEFT JOIN votes v ON v.selected_option_id = o.id
        WHERE o.poll_id = %s AND o.status = 1
        GROUP BY o.id, o.option
    """, (poll_id,))
    options     = cursor.fetchall()
    total_votes = sum(o['vote_count'] for o in options)

    options_data = []
    for o in options:
        percentage = round(
            (o['vote_count'] / total_votes * 100), 1
        ) if total_votes > 0 else 0
        options_data.append({
            'id':         o['id'],
            'text':       o['option'],
            'votes':      o['vote_count'],
            'percentage': percentage
        })

    cursor.execute("""
        SELECT
            v.id,
            v.created_at,
            vi.encrypted_identifier
        FROM votes v
        LEFT JOIN vote_identity vi ON vi.vote_id = v.id
        WHERE v.poll_id = %s
        ORDER BY v.created_at DESC
    """, (poll_id,))
    logs = cursor.fetchall()

    now = datetime.now(timezone.utc)

    end_time = poll['end_time']

    if end_time and end_time.tzinfo is None:
        end_time = end_time.replace(tzinfo=timezone.utc)

    is_expired = now > end_time

    cursor.close()
    conn.close()

    return render_template('poll_detail.html',
                           active_page  = 'polls',
                           poll         = poll,
                           options_data = options_data,
                           total_votes  = total_votes,
                           logs         = logs,
                           is_expired   = is_expired,
                           is_creator = True)


# ── Edit Poll GET ─────────────────────────────────────────
@app.route('/dashboard/poll/<string:token>/edit')
def edit_poll(token):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM polls
        WHERE share_token = %s
        AND user_id = %s AND status = 1
    """, (token, session['user_id']))
    poll = cursor.fetchone()

    if not poll:
        cursor.close()
        conn.close()
        return render_template('404.html'), 404

    poll_id    = poll["id"]
    now = datetime.now(timezone.utc)

    end_time = poll['end_time']

    if end_time and end_time.tzinfo is None:
        end_time = end_time.replace(tzinfo=timezone.utc)

    is_expired = now > end_time
    is_started = now >= poll['start_time']

    if is_expired:
        cursor.close()
        conn.close()
        flash('Cannot edit expired poll.', 'warning')
        return redirect(url_for('poll_detail', token=token))

    cursor.execute("""
        SELECT
            o.id, o.option, o.media_id,
            m.file_path, m.file_type,
            m.original_name
        FROM options o
        LEFT JOIN media m ON m.id = o.media_id
        WHERE o.poll_id = %s AND o.status = 1
    """, (poll_id,))
    options = cursor.fetchall()

    cursor.execute("""
        SELECT COUNT(*) as count FROM votes
        WHERE poll_id = %s
    """, (poll_id,))
    vote_count = cursor.fetchone()['count']

    cursor.close()
    conn.close()

    options_data = [{
        'id':            o['id'],
        'text':          o['option'],
        'media_id':      o['media_id'],
        'file_path':     o['file_path'],
        'file_type':     o['file_type'],
        'original_name': o['original_name']
    } for o in options]

    return render_template('edit_poll.html',
                           active_page  = 'polls',
                           poll         = poll,
                           options_data = options_data,
                           vote_count   = vote_count,
                           is_started    = is_started)


# ── Edit Poll POST ────────────────────────────────────────
@app.route('/dashboard/poll/<string:token>/edit',
           methods=['POST'])
def edit_poll_submit(token):
    if 'user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    data      = request.get_json()
    question  = data.get("question", "").strip()
    end_time  = data.get("end_time", "")
    poll_type = data.get("poll_type", "single")
    options   = data.get("options", [])

    if not end_time:
        return jsonify({"error": "End time required."}), 400

    valid_options = [
        o for o in options
        if (
    o.get("text", "").strip() or
    o.get("file_data") or
    o.get("media_id")
)
    ]
    if len(valid_options) < 2:
        return jsonify({
            "error": "At least 2 options required."
        }), 400

    option_texts = [
    o.get("text", "").strip().lower()
    for o in valid_options
    if o.get("text", "").strip()
    ]

    if len(option_texts) != len(set(option_texts)):
        return jsonify({
        "error": "All options must be unique."
        }), 400
    
    try:
        ist =  pytz.timezone('Asia/Kolkata')

        end_dt = datetime.fromisoformat(end_time)
        if end_dt.tzinfo is None:
            end_dt = ist.localize(end_dt)

        end_dt = end_dt.astimezone(timezone.utc)

        if end_dt <= datetime.now(timezone.utc):
            return jsonify({
                "error": "End time must be in the future."
            }), 400
    except ValueError:
        return jsonify({"error": "Invalid end time."}), 400

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM polls
        WHERE share_token = %s
        AND user_id = %s AND status = 1
    """, (token, session['user_id']))
    poll = cursor.fetchone()

    
    if not poll:
        cursor.close()
        conn.close()
        return jsonify({"error": "Poll not found."}), 404

    is_started = datetime.now(timezone.utc) >= poll['start_time']
    poll_id = poll["id"]

    cursor.execute("""
        SELECT COUNT(*) as count FROM votes
        WHERE poll_id = %s
    """, (poll_id,))
    vote_count = cursor.fetchone()['count']

    try:
        if is_started:
        # Poll already started → only end_time editable
            cursor.execute("""
            UPDATE polls SET end_time=%s WHERE id=%s
        """, (end_dt, poll_id))
        # Skip all option updates
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({
              "message": "Poll end time updated successfully!"
            }), 200

        if vote_count == 0:
            cursor.execute("""
                UPDATE polls
                SET question=%s, end_time=%s, poll_type=%s
                WHERE id=%s
            """, (question, end_dt, poll_type, poll_id))
        else:
            cursor.execute("""
                UPDATE polls
                SET question=%s, end_time=%s
                WHERE id=%s
            """, (question, end_dt, poll_id))

        os.makedirs(UPLOADS_FOLDER, exist_ok=True)

        for opt in valid_options:
            option_id   = opt.get("id")
            text        = opt.get("text", "").strip()
            file_data   = opt.get("file_data")
            file_name   = opt.get("file_name")
            file_type   = opt.get("file_type")
            file_size   = opt.get("file_size", 0)
            remove_file = opt.get("remove_file", False)
            media_id    = opt.get("media_id")
            new_media_id = media_id

            if remove_file:
                new_media_id = None
            elif file_data and file_name:
                ext         = os.path.splitext(
                    file_name)[1].lower()
                unique_name = f"{uuid.uuid4()}{ext}"
                file_path   = os.path.join(
                    UPLOADS_FOLDER, unique_name
                )
                file_bytes  = base64.b64decode(file_data)
                with open(file_path, 'wb') as f:
                    f.write(file_bytes)

                cursor.execute("""
                    INSERT INTO media
                        (file_name, file_path, file_type,
                         file_size, original_name,
                         created_at, created_id, status)
                    VALUES (%s,%s,%s,%s,%s,NOW(),%s,1)
                    RETURNING id
                """, (
                    unique_name,
                    f"uploads/{unique_name}",
                    file_type, file_size, file_name,
                    session.get('user_id')
                ))
                new_media_id = cursor.fetchone()['id']

            if option_id:
                cursor.execute("""
                    UPDATE options
                    SET option=%s, media_id=%s
                    WHERE id=%s AND poll_id=%s
                """, (text, new_media_id,
                      option_id, poll_id))
            else:
                cursor.execute("""
                    INSERT INTO options
                        (poll_id, option, media_id,
                         status, created_at, created_id)
                    VALUES (%s,%s,%s,1,NOW(),%s)
                """, (
                    poll_id, text, new_media_id,
                    session.get('user_id')
                ))

        kept_ids = [
            o.get("id") for o in valid_options
            if o.get("id")
        ]

        if kept_ids:
            placeholders = ",".join(
                ["%s"] * len(kept_ids)
            )
            cursor.execute(f"""
                UPDATE options SET status = 0
                WHERE poll_id = %s
                AND id NOT IN ({placeholders})
            """, [poll_id] + kept_ids)
        else:
            cursor.execute("""
                UPDATE options SET status = 0
                WHERE poll_id = %s
            """, (poll_id,))

        conn.commit()

    except Exception as e:
        conn.rollback()
        print(f"Edit poll error: {e}")
        return jsonify({
            "error": "Failed to update poll."
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": "Poll updated successfully!"
    }), 200


# ── Delete Poll ───────────────────────────────────────────
@app.route('/dashboard/poll/<string:token>/delete',
           methods=['POST'])
def delete_poll(token):
    if 'user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM polls
        WHERE share_token = %s AND user_id = %s
    """, (token, session['user_id']))
    poll = cursor.fetchone()

    if not poll:
        cursor.close()
        conn.close()
        return jsonify({"error": "Poll not found"}), 404

    cursor.execute(
        "UPDATE polls SET status = 0 WHERE id = %s",
        (poll["id"],)
    )
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({"message": "Poll deleted"}), 200


# ── Public Polls ──────────────────────────────────────────
@app.route('/polls')
def polls_list():
    page     = request.args.get('page', 1, type=int)
    per_page = 8
    offset   = (page - 1) * per_page

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) as count FROM polls
        WHERE status = 1
    """)
    total_count = cursor.fetchone()['count']

    cursor.execute("""
        SELECT
            p.id, p.question,
            p.start_time, p.end_time,
            p.share_token,
            u.first_name, u.last_name,
            COUNT(v.id) as vote_count,
            CASE
                WHEN p.end_time <= NOW()
                    THEN 'Expired'
                WHEN p.start_time > NOW()
                    THEN 'Not Started'
                ELSE 'Active'
            END as status
        FROM polls p
        LEFT JOIN votes v ON v.poll_id = p.id
        LEFT JOIN users u ON u.id = p.user_id
        WHERE p.status = 1
        GROUP BY p.id, p.question, p.start_time,
                 p.end_time, p.share_token,
                 u.first_name, u.last_name
        ORDER BY p.created_at DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    polls = cursor.fetchall()

    is_admin = session.get("role") == "admin"
    voted_poll_ids = []
    if "user_id" in session:
        cursor.execute("""
            SELECT DISTINCT poll_id
            FROM votes
            WHERE created_id = %s
        """, (session["user_id"],))

        voted = cursor.fetchall()

    # IMPORTANT: convert to list
        voted_poll_ids = [int(v["poll_id"]) for v in voted]
    cursor.close()
    conn.close()

    total_pages = (total_count + per_page - 1) // per_page

    return render_template('polls.html',
                           polls        = polls,
                           current_page = page,
                           total_pages  = total_pages,
                           total_count  = total_count,
                           is_admin=is_admin,
                            voted_poll_ids=voted_poll_ids)


# ── Login Validation ──────────────────────────────────────
@app.route('/login_validation', methods=['POST'])
def login_validation():
    email    = request.form.get('email')
    password = request.form.get('password')

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM users WHERE email = %s", (email,)
    )
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    if not user:
        flash('Invalid email or password.', 'danger')
        return redirect(url_for('login'))

    if user['status'] == 0:
        flash('Your account has been suspended.', 'danger')
        return redirect(url_for('login'))

    if not bcrypt.checkpw(
        password.encode('utf-8'),
        user['password'].encode('utf-8')
    ):
        flash('Invalid email or password.', 'danger')
        return redirect(url_for('login'))

    session['user_id']   = user['id']
    session['user_name'] = user['first_name']
    session['role']      = user['role']
    session.permanent    = True

    if user['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('poll.index'))


# ── Signup ────────────────────────────────────────────────
@app.route('/signup')
def signup():
    return render_template('signUp.html',
                           hide_navbar=True,
                           hide_footer=True)


# ── Add User ──────────────────────────────────────────────
@app.route('/add_user', methods=['POST'])
def add_user():
    fname    = request.form.get('fname')
    lname    = request.form.get('lname')
    email    = request.form.get('email')
    password = request.form.get('password')

    hashed_password = bcrypt.hashpw(
        password.encode('utf-8'),
        bcrypt.gensalt()
    ).decode('utf-8')

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM users WHERE email = %s", (email,)
    )
    existing = cursor.fetchone()

    if existing:
        cursor.close()
        conn.close()
        flash('An account with this email already exists.',
              'warning')
        return redirect(url_for('login'))

    cursor.execute("""
        INSERT INTO users
            (first_name, last_name, email, password,
             created_at, status)
        VALUES (%s,%s,%s,%s,NOW(),1)
        RETURNING id
    """, (fname, lname, email, hashed_password))
    new_id = cursor.fetchone()['id']

    cursor.execute(
        "UPDATE users SET created_id = %s WHERE id = %s",
        (new_id, new_id)
    )
    conn.commit()
    cursor.close()
    conn.close()

    flash('Account created! Please log in.', 'success')
    return redirect(url_for('login'))


# ── Admin Dashboard ───────────────────────────────────────
@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    conn        = get_db_connection()
    cursor      = conn.cursor()
    filter_user = request.args.get('user_id', None)

    cursor.execute("""
        SELECT COUNT(*) as count FROM users
        WHERE role = 'user' AND status = 1
    """)
    total_users = cursor.fetchone()['count']

    if filter_user:
        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE user_id = %s AND status = 1
        """, (filter_user,))
        total_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM votes
            WHERE poll_id IN (
                SELECT id FROM polls WHERE user_id = %s
            )
        """, (filter_user,))
        total_votes = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE user_id = %s AND status = 1
            AND end_time > NOW()
        """, (filter_user,))
        active_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT
                p.id, p.share_token, p.question,
                p.end_time, p.start_time,
                u.first_name, u.last_name,
                COUNT(v.id) as vote_count,
                CASE
                    WHEN p.end_time <= NOW()
                        THEN 'Expired'
                    WHEN p.start_time > NOW()
                        THEN 'Not Started'
                    ELSE 'Active'
                END as status
            FROM polls p
            LEFT JOIN votes v ON v.poll_id = p.id
            LEFT JOIN users u ON u.id = p.user_id
            WHERE p.status = 1 AND p.user_id = %s
            GROUP BY p.id, p.share_token, p.question,
                     p.end_time, p.start_time,
                     u.first_name, u.last_name
            ORDER BY p.created_at DESC
            LIMIT 5
        """, (filter_user,))
        recent_polls = cursor.fetchall()

    else:
        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
        """)
        total_polls = cursor.fetchone()['count']

        cursor.execute(
            "SELECT COUNT(*) as count FROM votes"
        )
        total_votes = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1 AND end_time > NOW()
        """)
        active_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT
                p.id, p.share_token, p.question,
                p.end_time, p.start_time,
                u.first_name, u.last_name,
                COUNT(v.id) as vote_count,
                CASE
                    WHEN p.end_time <= NOW()
                        THEN 'Expired'
                    WHEN p.start_time > NOW()
                        THEN 'Not Started'
                    ELSE 'Active'
                END as status
            FROM polls p
            LEFT JOIN votes v ON v.poll_id = p.id
            LEFT JOIN users u ON u.id = p.user_id
            WHERE p.status = 1
            GROUP BY p.id, p.share_token, p.question,
                     p.end_time, p.start_time,
                     u.first_name, u.last_name
            ORDER BY p.created_at DESC
            LIMIT 5
        """)
        recent_polls = cursor.fetchall()

    cursor.execute("""
        SELECT id, first_name, last_name, email
        FROM users
        WHERE role = 'user' AND status = 1
        ORDER BY first_name
    """)
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('admin_dashboard.html',
                           active_page  = 'admin_dashboard',
                           total_users  = total_users,
                           total_polls  = total_polls,
                           total_votes  = total_votes,
                           active_polls = active_polls,
                           recent_polls = recent_polls,
                           users        = users,
                           filter_user  = filter_user)


# ── Admin Polls ───────────────────────────────────────────
@app.route('/admin/polls')
@admin_required
def admin_polls():
    conn   = get_db_connection()
    cursor = conn.cursor()

    page     = request.args.get('page', 1, type=int)
    per_page = 10
    offset   = (page - 1) * per_page

    filter_user   = request.args.get('user_id', '')
    filter_status = request.args.get('status', '')
    filter_date   = request.args.get('date', '')

    where_clauses = ["p.status = 1"]
    params        = []

    if filter_user:
        where_clauses.append("p.user_id = %s")
        params.append(filter_user)

    if filter_status == 'active':
        where_clauses.append(
            "p.end_time > NOW() AND p.start_time <= NOW()"
        )
    elif filter_status == 'expired':
        where_clauses.append("p.end_time <= NOW()")
    elif filter_status == 'not_started':
        where_clauses.append("p.start_time > NOW()")

    if filter_date:
        where_clauses.append(
            "DATE(p.created_at) = %s"
        )
        params.append(filter_date)

    where_sql = " AND ".join(where_clauses)

    cursor.execute(f"""
        SELECT COUNT(*) as count FROM polls p
        WHERE {where_sql}
    """, params)
    total_count = cursor.fetchone()['count']

    cursor.execute(f"""
        SELECT
            p.id, p.share_token, p.question,
            p.start_time, p.end_time,
            p.poll_type,
            u.first_name, u.last_name,
            COUNT(v.id) as vote_count,
            CASE
                WHEN p.end_time <= NOW()
                    THEN 'Expired'
                WHEN p.start_time > NOW()
                    THEN 'Not Started'
                ELSE 'Active'
            END as status
        FROM polls p
        LEFT JOIN votes v ON v.poll_id = p.id
        LEFT JOIN users u ON u.id = p.user_id
        WHERE {where_sql}
        GROUP BY p.id, p.share_token, p.question,
                 p.start_time, p.end_time, p.poll_type,
                 u.first_name, u.last_name
        ORDER BY p.created_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    polls = cursor.fetchall()

    cursor.execute("""
        SELECT id, first_name, last_name
        FROM users WHERE role = 'user' AND status = 1
        ORDER BY first_name
    """)
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    total_pages = (total_count + per_page - 1) // per_page

    return render_template('admin_polls.html',
                           active_page   = 'admin_polls',
                           polls         = polls,
                           users         = users,
                           current_page  = page,
                           total_pages   = total_pages,
                           total_count   = total_count,
                           filter_user   = filter_user,
                           filter_status = filter_status,
                           filter_date   = filter_date)


# ── Admin Users ───────────────────────────────────────────
@app.route('/admin/users')
@admin_required
def admin_users():
    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            u.id, u.first_name, u.last_name,
            u.email, u.status, u.created_at,
            COUNT(p.id) as poll_count
        FROM users u
        LEFT JOIN polls p ON p.user_id = u.id
                          AND p.status = 1
        WHERE u.role = 'user'
        GROUP BY u.id, u.first_name, u.last_name,
                 u.email, u.status, u.created_at
        ORDER BY u.created_at DESC
    """)
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('admin_users.html',
                           active_page = 'admin_users',
                           users       = users)


# ── Admin Edit User ───────────────────────────────────────
@app.route('/admin/users/<int:user_id>/edit',
           methods=['POST'])
@admin_required
def admin_edit_user(user_id):
    data       = request.get_json()
    first_name = data.get('first_name', '').strip()
    last_name  = data.get('last_name', '').strip()
    email      = data.get('email', '').strip()
    password   = data.get('password', '').strip()
    status     = data.get('status', 1)

    if not first_name or not last_name or not email:
        return jsonify({
            "error": "Name and email are required."
        }), 400

    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM users WHERE id = %s AND role='user'",
        (user_id,)
    )
    user = cursor.fetchone()

    if not user:
        cursor.close()
        conn.close()
        return jsonify({"error": "User not found."}), 404

    cursor.execute("""
        SELECT id FROM users
        WHERE email = %s AND id != %s
    """, (email, user_id))
    if cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({
            "error": "Email already in use."
        }), 400

    try:
        if password:
            hashed = bcrypt.hashpw(
                password.encode('utf-8'),
                bcrypt.gensalt()
            ).decode('utf-8')
            cursor.execute("""
                UPDATE users
                SET first_name=%s, last_name=%s,
                    email=%s, password=%s, status=%s
                WHERE id=%s
            """, (first_name, last_name,
                  email, hashed, status, user_id))
        else:
            cursor.execute("""
                UPDATE users
                SET first_name=%s, last_name=%s,
                    email=%s, status=%s
                WHERE id=%s
            """, (first_name, last_name,
                  email, status, user_id))

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "error": "Failed to update user."
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": "User updated successfully!"
    }), 200


# ── Admin Ban User ────────────────────────────────────────
@app.route('/admin/users/<int:user_id>/ban',
           methods=['POST'])
@admin_required
def admin_ban_user(user_id):
    conn   = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM users WHERE id=%s AND role='user'",
        (user_id,)
    )
    user = cursor.fetchone()

    if not user:
        cursor.close()
        conn.close()
        return jsonify({"error": "User not found."}), 404

    new_status = 0 if user['status'] == 1 else 1
    action     = "banned" if new_status == 0 else "unbanned"

    cursor.execute(
        "UPDATE users SET status=%s WHERE id=%s",
        (new_status, user_id)
    )
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({
        "message":    f"User {action} successfully!",
        "new_status": new_status
    }), 200


# ── Admin Reports ─────────────────────────────────────────
@app.route('/admin/reports')
@admin_required
def admin_reports():
    report_type   = request.args.get('type', '')
    filter_status = request.args.get('status', '')

    conn   = get_db_connection()
    cursor = conn.cursor()
    report_data = {}

    if report_type == 'users':
        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
        """)
        total_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user' AND status = 1
        """)
        active_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user' AND status = 0
        """)
        inactive_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
            AND DATE_TRUNC('month', created_at) =
                DATE_TRUNC('month', NOW())
        """)
        new_this_month = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
            AND DATE(created_at) = CURRENT_DATE
        """)
        new_today = cursor.fetchone()['count']

        status_filter = ""
        if filter_status == 'active':
            status_filter = "AND u.status = 1"
        elif filter_status == 'inactive':
            status_filter = "AND u.status = 0"

        cursor.execute(f"""
            SELECT
                u.id, u.first_name, u.last_name,
                u.email, u.status, u.created_at,
                COUNT(p.id) as poll_count
            FROM users u
            LEFT JOIN polls p ON p.user_id = u.id
                              AND p.status = 1
            WHERE u.role = 'user'
            {status_filter}
            GROUP BY u.id, u.first_name, u.last_name,
                     u.email, u.status, u.created_at
            ORDER BY u.created_at DESC
        """)
        users_list = cursor.fetchall()

        report_data = {
            'total_users':    total_users,
            'active_users':   active_users,
            'inactive_users': inactive_users,
            'new_this_month': new_this_month,
            'new_today':      new_today,
            'users_list':     users_list
        }

    elif report_type == 'polls':
        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
        """)
        total_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
            AND end_time > NOW()
            AND start_time <= NOW()
        """)
        active_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1 AND end_time <= NOW()
        """)
        expired_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1 AND start_time > NOW()
        """)
        not_started_polls = cursor.fetchone()['count']

        cursor.execute(
            "SELECT COUNT(*) as count FROM votes"
        )
        total_votes = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
            AND DATE_TRUNC('month', created_at) =
                DATE_TRUNC('month', NOW())
        """)
        polls_this_month = cursor.fetchone()['count']

        if filter_status == 'active':
            where = """
                p.status = 1
                AND p.end_time > NOW()
                AND p.start_time <= NOW()
            """
        elif filter_status == 'expired':
            where = "p.status = 1 AND p.end_time <= NOW()"
        elif filter_status == 'not_started':
            where = "p.status = 1 AND p.start_time > NOW()"
        else:
            where = "p.status = 1"

        cursor.execute(f"""
            SELECT
                p.id, p.question,
                p.start_time, p.end_time,
                p.poll_type, p.share_token,
                u.first_name, u.last_name,
                COUNT(v.id) as vote_count,
                CASE
                    WHEN p.end_time <= NOW()
                        THEN 'Expired'
                    WHEN p.start_time > NOW()
                        THEN 'Not Started'
                    ELSE 'Active'
                END as poll_status
            FROM polls p
            LEFT JOIN votes v ON v.poll_id = p.id
            LEFT JOIN users u ON u.id = p.user_id
            WHERE {where}
            GROUP BY p.id, p.question, p.start_time,
                     p.end_time, p.poll_type,
                     p.share_token,
                     u.first_name, u.last_name
            ORDER BY p.created_at DESC
        """)
        polls_list = cursor.fetchall()

        report_data = {
            'total_polls':       total_polls,
            'active_polls':      active_polls,
            'expired_polls':     expired_polls,
            'not_started_polls': not_started_polls,
            'total_votes':       total_votes,
            'polls_this_month':  polls_this_month,
            'polls_list':        polls_list
        }

    cursor.close()
    conn.close()

    return render_template('admin_reports.html',
                           active_page   = 'admin_reports',
                           report_type   = report_type,
                           filter_status = filter_status,
                           report_data   = report_data)


def _get_report_data_for_export(report_type, filter_status):
    conn = get_db_connection()
    cursor = conn.cursor()
    report_data = {}

    if report_type == 'users':
        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
        """)
        total_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user' AND status = 1
        """)
        active_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user' AND status = 0
        """)
        inactive_users = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
            AND DATE_TRUNC('month', created_at) =
                DATE_TRUNC('month', NOW())
        """)
        new_this_month = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM users
            WHERE role = 'user'
            AND DATE(created_at) = CURRENT_DATE
        """)
        new_today = cursor.fetchone()['count']

        status_filter = ""
        if filter_status == 'active':
            status_filter = "AND u.status = 1"
        elif filter_status == 'inactive':
            status_filter = "AND u.status = 0"

        cursor.execute(f"""
            SELECT
                u.id, u.first_name, u.last_name,
                u.email, u.status, u.created_at,
                COUNT(p.id) as poll_count
            FROM users u
            LEFT JOIN polls p ON p.user_id = u.id
                              AND p.status = 1
            WHERE u.role = 'user'
            {status_filter}
            GROUP BY u.id, u.first_name, u.last_name,
                     u.email, u.status, u.created_at
            ORDER BY u.created_at DESC
        """)
        users_list = cursor.fetchall()

        report_data = {
            'total_users': total_users,
            'active_users': active_users,
            'inactive_users': inactive_users,
            'new_this_month': new_this_month,
            'new_today': new_today,
            'users_list': users_list
        }

    elif report_type == 'polls':
        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
        """)
        total_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
            AND end_time > NOW()
            AND start_time <= NOW()
        """)
        active_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1 AND end_time <= NOW()
        """)
        expired_polls = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1 AND start_time > NOW()
        """)
        not_started_polls = cursor.fetchone()['count']

        cursor.execute("SELECT COUNT(*) as count FROM votes")
        total_votes = cursor.fetchone()['count']

        cursor.execute("""
            SELECT COUNT(*) as count FROM polls
            WHERE status = 1
            AND DATE_TRUNC('month', created_at) =
                DATE_TRUNC('month', NOW())
        """)
        polls_this_month = cursor.fetchone()['count']

        if filter_status == 'active':
            where = """
                p.status = 1
                AND p.end_time > NOW()
                AND p.start_time <= NOW()
            """
        elif filter_status == 'expired':
            where = "p.status = 1 AND p.end_time <= NOW()"
        elif filter_status == 'not_started':
            where = "p.status = 1 AND p.start_time > NOW()"
        else:
            where = "p.status = 1"

        cursor.execute(f"""
            SELECT
                p.id, p.question,
                p.start_time, p.end_time,
                p.poll_type, p.share_token,
                u.first_name, u.last_name,
                COUNT(v.id) as vote_count,
                CASE
                    WHEN p.end_time <= NOW()
                        THEN 'Expired'
                    WHEN p.start_time > NOW()
                        THEN 'Not Started'
                    ELSE 'Active'
                END as poll_status
            FROM polls p
            LEFT JOIN votes v ON v.poll_id = p.id
            LEFT JOIN users u ON u.id = p.user_id
            WHERE {where}
            GROUP BY p.id, p.question, p.start_time,
                     p.end_time, p.poll_type,
                     p.share_token,
                     u.first_name, u.last_name
            ORDER BY p.created_at DESC
        """)
        polls_list = cursor.fetchall()

        report_data = {
            'total_polls': total_polls,
            'active_polls': active_polls,
            'expired_polls': expired_polls,
            'not_started_polls': not_started_polls,
            'total_votes': total_votes,
            'polls_this_month': polls_this_month,
            'polls_list': polls_list
        }

    cursor.close()
    conn.close()
    return report_data


def _autosize_columns(sheet):
    for col_idx, col in enumerate(sheet.iter_cols(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column
    ), start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            cell_value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(cell_value))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _build_excel_report(report_type, filter_status, report_data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F4E78")
    bold_font = Font(bold=True)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = "Users Summary Report" if report_type == "users" else "Polls Summary Report"
    filter_text = filter_status.replace("_", " ").title() if filter_status else "All"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = f"PulsePoll - {title}"
    sheet["A1"].font = title_font
    sheet["A2"] = "Generated At"
    sheet["B2"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    sheet["A3"] = "Filter"
    sheet["B3"] = filter_text
    sheet["A2"].font = bold_font
    sheet["A3"].font = bold_font

    row = 5
    if report_type == "users":
        summary = [
            ("Total Users", report_data.get("total_users", 0)),
            ("Active Users", report_data.get("active_users", 0)),
            ("Inactive Users", report_data.get("inactive_users", 0)),
            ("New This Month", report_data.get("new_this_month", 0)),
            ("New Today", report_data.get("new_today", 0)),
        ]
        headers = ["#", "Name", "Email", "Polls Created", "Joined Date", "Status"]
        rows = []
        for idx, user in enumerate(report_data.get("users_list", []), start=1):
            joined = user["created_at"].strftime("%d %b %Y") if user.get("created_at") else "-"
            full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
            status_text = "Active" if user.get("status") == 1 else "Inactive"
            rows.append([idx, full_name, user.get("email", ""), user.get("poll_count", 0), joined, status_text])
    else:
        summary = [
            ("Total Polls", report_data.get("total_polls", 0)),
            ("Active Polls", report_data.get("active_polls", 0)),
            ("Expired Polls", report_data.get("expired_polls", 0)),
            ("Not Started", report_data.get("not_started_polls", 0)),
            ("Total Votes", report_data.get("total_votes", 0)),
            ("Created This Month", report_data.get("polls_this_month", 0)),
        ]
        headers = ["#", "Question", "Created By", "Type", "Votes", "End Time", "Status"]
        rows = []
        for idx, poll in enumerate(report_data.get("polls_list", []), start=1):
            creator = f"{poll.get('first_name', '')} {poll.get('last_name', '')}".strip()
            end_time = poll["end_time"].strftime("%d %b %Y, %I:%M %p") if poll.get("end_time") else "-"
            rows.append([idx, poll.get("question", ""), creator, poll.get("poll_type", "").title(),
                         poll.get("vote_count", 0), end_time, poll.get("poll_status", "")])

    sheet[f"A{row}"] = "Summary"
    sheet[f"A{row}"].font = Font(size=12, bold=True, color="1F4E78")
    row += 1
    for label, value in summary:
        sheet[f"A{row}"] = label
        sheet[f"A{row}"].font = bold_font
        sheet[f"B{row}"] = value
        row += 1

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_data in rows:
        row += 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col_idx in (1, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

    _autosize_columns(sheet)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _build_pdf_report(report_type, filter_status, report_data):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except Exception:
        return None

    title = "Users Summary Report" if report_type == "users" else "Polls Summary Report"
    filter_text = filter_status.replace("_", " ").title() if filter_status else "All"

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>PulsePoll - {title}</b>", styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Generated At: {datetime.now().strftime('%d %b %Y, %I:%M %p')}", styles["Normal"]))
    story.append(Paragraph(f"Filter: {filter_text}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if report_type == "users":
        summary_rows = [
            ["Total Users", report_data.get("total_users", 0)],
            ["Active Users", report_data.get("active_users", 0)],
            ["Inactive Users", report_data.get("inactive_users", 0)],
            ["New This Month", report_data.get("new_this_month", 0)],
            ["New Today", report_data.get("new_today", 0)],
        ]
        details = [["#", "Name", "Email", "Polls Created", "Joined Date", "Status"]]
        for idx, user in enumerate(report_data.get("users_list", []), start=1):
            joined = user["created_at"].strftime("%d %b %Y") if user.get("created_at") else "-"
            full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
            status_text = "Active" if user.get("status") == 1 else "Inactive"
            details.append([idx, full_name, user.get("email", ""), user.get("poll_count", 0), joined, status_text])
    else:
        summary_rows = [
            ["Total Polls", report_data.get("total_polls", 0)],
            ["Active Polls", report_data.get("active_polls", 0)],
            ["Expired Polls", report_data.get("expired_polls", 0)],
            ["Not Started", report_data.get("not_started_polls", 0)],
            ["Total Votes", report_data.get("total_votes", 0)],
            ["Created This Month", report_data.get("polls_this_month", 0)],
        ]
        details = [["#", "Question", "Created By", "Type", "Votes", "End Time", "Status"]]
        for idx, poll in enumerate(report_data.get("polls_list", []), start=1):
            creator = f"{poll.get('first_name', '')} {poll.get('last_name', '')}".strip()
            end_time = poll["end_time"].strftime("%d %b %Y, %I:%M %p") if poll.get("end_time") else "-"
            details.append([idx, poll.get("question", ""), creator, poll.get("poll_type", "").title(),
                            poll.get("vote_count", 0), end_time, poll.get("poll_status", "")])

    summary_table = Table([["Summary", "Value"]] + summary_rows, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    detail_table = Table(details, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(detail_table)

    doc.build(story)
    output.seek(0)
    return output


@app.route('/admin/reports/export')
@admin_required
def export_admin_report():
    report_type = request.args.get('type', '')
    filter_status = request.args.get('status', '')
    export_format = request.args.get('format', 'pdf').lower()

    if report_type not in ('users', 'polls'):
        return jsonify({"error": "Invalid report type."}), 400

    if export_format not in ('pdf', 'xlsx'):
        return jsonify({"error": "Invalid export format."}), 400

    report_data = _get_report_data_for_export(report_type, filter_status)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_type}_report_{ts}"

    if export_format == "xlsx":
        excel_stream = _build_excel_report(report_type, filter_status, report_data)
        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=f"{filename}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    pdf_stream = _build_pdf_report(report_type, filter_status, report_data)
    if pdf_stream is None:
        return jsonify({
            "error": "PDF export requires reportlab package. Install it and retry."
        }), 500

    return send_file(
        pdf_stream,
        as_attachment=True,
        download_name=f"{filename}.pdf",
        mimetype="application/pdf"
    )


# ── Logout ────────────────────────────────────────────────
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('user_name', None)
    session.pop('role', None)
    return redirect(url_for('poll.index'))


# ── Error Handlers ────────────────────────────────────────
@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(500)
def internal_error(e):
    return render_template("500.html"), 500


# ── Run ───────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True)

